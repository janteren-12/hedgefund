"""
Builds a single downloadable Excel workbook out of the same data the web
pages show, reusing the queries.py functions so there's only one place
that computes each view.

Sheet layout:
  - One sheet per fund's current positions (mirrors the collapsible
    sections on the Current Positions page)
  - "Reduced or Exited" / "New Buys & Increases" (mirrors Selling &
    Buying Activity, with a Fund column since those tables already span
    every fund)
  - "Overlap" (mirrors the stock x fund matrix)
  - "By Strategy" (mirrors the strategy groupings)
"""

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font

import queries

BOLD = Font(bold=True)
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")


def _safe_sheet_name(name, used_names):
    cleaned = _INVALID_SHEET_CHARS.sub(" ", name).strip()[:31] or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used_names:
        candidate = f"{cleaned[: 31 - len(str(suffix)) - 1]} {suffix}"
        suffix += 1
    used_names.add(candidate.lower())
    return candidate


def _write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = BOLD
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 45)


def build_workbook(conn):
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    for fund in queries.get_positions_by_fund(conn):
        sheet_name = _safe_sheet_name(fund["fund_name"], used_names)
        ws = wb.create_sheet(sheet_name)
        rows = [
            (r["ticker"], r["company_name"], r["value_usd"], round(r["pct"], 2))
            for r in fund["rows"]
        ]
        _write_table(ws, ["Ticker", "Company", "Market Value ($)", "% of Portfolio"], rows)

    selling = queries.get_selling_buying_by_fund(conn)

    ws = wb.create_sheet(_safe_sheet_name("Reduced or Exited", used_names))
    rows = [
        (
            fund["fund_name"],
            r["ticker"],
            r["company_name"],
            r["shares_previous"],
            r["shares_latest"],
            round(r["pct_change"], 1) if r["pct_change"] is not None else None,
            "SOLD OUT" if r["sold_out"] else "",
        )
        for fund in selling
        for r in fund["reduced"]
    ]
    _write_table(
        ws,
        ["Fund", "Ticker", "Company", "Shares (prev qtr)", "Shares (this qtr)", "% Change", ""],
        rows,
    )

    ws = wb.create_sheet(_safe_sheet_name("New Buys and Increases", used_names))
    rows = [
        (
            fund["fund_name"],
            r["ticker"],
            r["company_name"],
            r["shares_previous"],
            r["shares_latest"],
            round(r["pct_change"], 1) if r["pct_change"] is not None else None,
            "NEW" if r["is_new"] else "",
        )
        for fund in selling
        for r in fund["increased"]
    ]
    _write_table(
        ws,
        ["Fund", "Ticker", "Company", "Shares (prev qtr)", "Shares (this qtr)", "% Change", ""],
        rows,
    )

    fund_columns, overlap_rows = queries.get_overlap(conn)
    ws = wb.create_sheet(_safe_sheet_name("Overlap", used_names))
    headers = ["Ticker", "Company", "# Funds", "Weighted Avg %"] + [f["name"] for f in fund_columns]
    rows = [
        (
            r["ticker"],
            r["company_name"],
            r["fund_count"],
            round(r["weighted_avg_pct"], 2),
            *(
                round(r["pct_by_fund"][f["name"]], 2) if f["name"] in r["pct_by_fund"] else None
                for f in fund_columns
            ),
        )
        for r in overlap_rows
    ]
    _write_table(ws, headers, rows)

    ws = wb.create_sheet(_safe_sheet_name("By Strategy", used_names))
    rows = [
        (group["focus"], fund["name"], fund["known_for"], fund["total_value"], fund["holding_count"], fund["period"])
        for group in queries.get_funds_by_focus(conn)
        for fund in group["funds"]
    ]
    _write_table(
        ws,
        ["Focus", "Fund", "Known For", "13F Portfolio Value ($)", "# Holdings", "As Of"],
        rows,
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
